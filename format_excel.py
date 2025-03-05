import openpyxl
from openpyxl import Workbook

# Load the source Excel file
source_wb = openpyxl.load_workbook('fresher.xlsx.xlsx')
source_ws = source_wb.active

# Create a new workbook with the correct structure
new_wb = Workbook()
new_ws = new_wb.active

# Set the headers
headers = ['Job Title', 'Date of application', 'Name', 'Email ID', 'Phone Number', 'Current Location', 'Preferred Locations', 'Total Experience', 'Curr. Company name', 'Curr. Company Designation', 'Department', 'Role', 'Industry', 'Key Skills', 'Annual Salary', 'Notice period/ Availability to join', 'Resume Headline', 'Summary', 'Under Graduation degree', 'UG Specialization', 'UG University/institute Name', 'UG Graduation year', 'Post graduation degree', 'PG specialization', 'PG university/institute name', 'PG graduation year', 'Doctorate degree', 'Doctorate specialization', 'Doctorate university/institute name', 'Doctorate graduation year', 'Gender', 'Marital Status', 'Home Town/City', 'Pin Code', 'Work permit for USA', 'Date of Birth', 'Permanent Address',]
for col, header in enumerate(headers, 1):
    new_ws.cell(row=1, column=col, value=header)

# Copy data from source to new workbook
# We'll try to match columns based on similar headers
source_headers = {}
for cell in source_ws[1]:
    if cell.value:
        header = str(cell.value).lower().strip()
        col_idx = cell.column
        source_headers[header] = col_idx

# Map source columns to target columns
column_mapping = {
    'Job Title': ['Job Title'],
    'Date of application': ['Date of application'],
    'Name': ['Name'],
    'Email ID': ['Email ID'],
    'Phone Number': ['Phone Number'],
    'Current Location': ['Current Location'],
    'Preferred Locations': ['Preferred Locations'],
    'Total Experience': ['Total Experience'],
    'Curr. Company name': ['Curr. Company name'],
    'Curr. Company Designation': ['Curr. Company Designation'],
    'Department': ['Department'],
    'Role': ['Role'],
    'Industry': ['Industry'],
    'Key Skills': ['Key Skills'],
    'Annual Salary': ['Annual Salary'],
    'Notice period/ Availability to join': ['Notice period/ Availability to join'],
    'Resume Headline': ['Resume Headline'],
    'Summary': ['Summary'],
    'Under Graduation degree': ['Under Graduation degree'],
    'UG Specialization': ['UG Specialization'],
    'UG University/institute Name': ['UG University/institute Name'],
    'UG Graduation year': ['UG Graduation year'],
    'Post graduation degree': ['Post graduation degree'],
    'PG specialization': ['PG specialization'],
    'PG university/institute name': ['PG university/institute name'],
    'PG graduation year': ['PG graduation year'],
    'Doctorate degree': ['Doctorate degree'],
    'Doctorate specialization': ['Doctorate specialization'],
    'Doctorate university/institute name': ['Doctorate university/institute name'],
    'Doctorate graduation year': ['Doctorate graduation year'],
    'Gender': ['Gender'],
    'Marital Status': ['Marital Status'],
    'Home Town/City': ['Home Town/City'],
    'Pin Code': ['Pin Code'],
    'Work permit for USA': ['Work permit for USA'],
    'Date of Birth': ['Date of Birth'],
    'Permanent Address': ['Permanent Address'],
}

# Find column indices in source file
source_indices = {}
for target_col, possible_headers in column_mapping.items():
    for header, idx in source_headers.items():
        if header in possible_headers:
            source_indices[target_col] = idx
            break

# Copy data
row_idx = 2
for row in list(source_ws.rows)[1:]:  # Skip header row
    # Only process rows that have at least one non-empty cell
    if any(cell.value for cell in row):
        for col_idx, header in enumerate(headers, 1):
            target_cell = new_ws.cell(row=row_idx, column=col_idx)
            header_lower = header.lower()
            if header_lower in source_indices:
                source_value = row[source_indices[header_lower] - 1].value
                target_cell.value = str(source_value) if source_value is not None else ""
        row_idx += 1

# Save the formatted file
new_wb.save('fresher.xlsx')
print("Excel file has been formatted successfully!")

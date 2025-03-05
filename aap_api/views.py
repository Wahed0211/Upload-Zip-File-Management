from django.shortcuts import render, redirect
from rest_framework import viewsets, status
from rest_framework.decorators import action, api_view
from rest_framework.response import Response
from rest_framework.permissions import AllowAny
from rest_framework.authentication import SessionAuthentication, BasicAuthentication
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from django.http import HttpResponse, JsonResponse
from .models import ExcelData, UploadedExcel
from .serializers import ExcelDataSerializer
import openpyxl
from django.http import HttpResponse
from datetime import datetime
from django.views.generic import ListView
import zipfile
import io
import os
from rest_framework.parsers import MultiPartParser, FormParser
import traceback
from .models import DownloadHistory
from django.core.files.storage import FileSystemStorage
from .tasks import process_excel_file, process_zip_file
import tempfile
import logging
from django.contrib.auth.mixins import LoginRequiredMixin 
import pandas as pd
from django.conf import settings
from .resource import ExcelDataResource
from django.contrib import messages
from tablib import Dataset


logger = logging.getLogger(__name__)

class ExcelDataListView(LoginRequiredMixin, ListView):
    model = ExcelData
    template_name = 'aap_api/item_list.html'
    context_object_name = 'items'
    paginate_by = 20  # Increased from 10 for better performance
    login_url = '/api/auth/login/'  # Redirect to login page if not authenticated

    def get_queryset(self):
        queryset = ExcelData.objects.select_related().order_by('-date_of_application')
        
        # Get all filter parameters at once
        filters = {}
        if job_title := self.request.GET.get('job_title'):
            filters['job_title'] = job_title
        if location := self.request.GET.get('location'):
            filters['current_location__icontains'] = location
        if company := self.request.GET.get('company'):
            filters['current_company_name__icontains'] = company
        if experience := self.request.GET.get('experience'):
            filters['total_experience__icontains'] = experience
        if name_search := self.request.GET.get('name_search'):
            filters['name__icontains'] = name_search

        return queryset.filter(**filters)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Add filter choices to context
        context['job_titles'] = ExcelData.JOB_TITLE_CHOICES
        context['locations'] = ExcelData.objects.values_list('current_location', flat=True).distinct()
        context['companies'] = ExcelData.objects.values_list('current_company_name', flat=True).distinct()
        
        # Add current filter values to context
        context['current_filters'] = {
            'job_title': self.request.GET.get('job_title', ''),
            'location': self.request.GET.get('location', ''),
            'company': self.request.GET.get('company', ''),
            'experience': self.request.GET.get('experience', ''),
            'name_search': self.request.GET.get('name_search', ''),
        }
        
        # Add total count
        context['total_records'] = self.get_queryset().count()
        
        return context

    def post(self, request, *args, **kwargs):
        if 'delete_record' in request.POST:
            record_id = request.POST.get('record_id')
            try:
                record = ExcelData.objects.get(id=record_id)
                record.delete()
                return JsonResponse({'status': 'success'})
            except ExcelData.DoesNotExist:
                return JsonResponse({'status': 'error', 'message': 'Record not found'})
        return redirect('aap_api:item_list')

@method_decorator(csrf_exempt, name='dispatch')
class ExcelDataViewSet(viewsets.ModelViewSet):
    """
    ViewSet for viewing and editing ExcelData instances.
    Provides default list, create, retrieve, update and delete actions.
    Additional actions:
    - import_excel: POST endpoint to import data from Excel file
    - export_excel: GET endpoint to export data to Excel file
    - import_zip_excel: POST endpoint to import data from ZIP file containing multiple Excel files
    """
    queryset = ExcelData.objects.select_related().all()
    serializer_class = ExcelDataSerializer
    permission_classes = [AllowAny]
    parser_classes = (MultiPartParser, FormParser)
    authentication_classes = []  # Remove authentication for testing
    page_size = 50  # Add pagination

    def get_queryset(self):
        queryset = ExcelData.objects.select_related().all()
        
        # Get all filter parameters at once
        filters = {}
        if job_title := self.request.query_params.get('job_title'):
            filters['job_title'] = job_title
        if location := self.request.query_params.get('location'):
            filters['current_location__icontains'] = location
        if company := self.request.query_params.get('company'):
            filters['current_company_name__icontains'] = company
            
        # Apply all filters at once
        if filters:
            queryset = queryset.filter(**filters)
            
        return queryset.order_by('-date_of_application')

    def get_column_indices(self, worksheet):
        """Helper function to find column indices based on headers"""
        headers = {}
        header_mappings = {
            # Standard format
            'job_title': ['job_title', 'job title'],
            'date_of_application': ['date_of_application', 'date of application'],
            'email_id': ['email_id', 'email id'],
            'phone_number': ['phone_number', 'phone number'],
            'current_location': ['current_location', 'current location'],
            'preferred_locations': ['preferred_locations', 'preferred locations'],
            'total_experience': ['total_experience', 'total experience'],
            'current_company_name': ['current_company_name', 'current company name', 'curr. company name'],
            'current_company_designation': ['current_company_designation', 'current company designation', 'curr. company designation'],
            'key_skills': ['key_skills', 'key skills'],
            'annual_salary': ['annual_salary', 'annual salary'],
            'notice_period_availability_to_join': ['notice_period_availability_to_join', 'notice period availability to join', 'notice period/ availability to join'],
            'resume_headline': ['resume_headline', 'resume headline'],
            'under_graduation_degree': ['under_graduation_degree', 'under graduation degree'],
            'ug_specialization': ['ug_specialization', 'ug specialization'],
            'ug_university_institute_name': ['ug_university_institute_name', 'ug university/institute name'],
            'ug_graduation_year': ['ug_graduation_year', 'ug graduation year'],
            'post_graduation_degree': ['post_graduation_degree', 'post graduation degree'],
            'pg_specialization': ['pg_specialization', 'pg specialization'],
            'pg_university_institute_name': ['pg_university_institute_name', 'pg university/institute name'],
            'pg_graduation_year': ['pg_graduation_year', 'pg graduation year'],
            'doctorate_degree': ['doctorate_degree', 'doctorate degree'],
            'doctorate_specialization': ['doctorate_specialization', 'doctorate specialization']
        }

        for cell in worksheet[1]:  # First row
            if cell.value:
                header = str(cell.value).lower().strip()
                # Find the matching standard header
                for standard_header, variations in header_mappings.items():
                    if header in variations:
                        headers[standard_header] = cell.column - 1
                        break
                else:
                    # If no mapping found, use the header as is
                    headers[header] = cell.column - 1
        return headers

    def parse_date(self, value):
        """Helper function to parse dates from Excel"""
        if not value:
            return None
        if isinstance(value, datetime):
            return value.date()
        try:
            # Try different date formats
            for fmt in ['%Y-%m-%d', '%Y-%m-%d %H:%M:%S', '%d-%m-%Y', '%m/%d/%Y']:
                try:
                    return datetime.strptime(str(value).split()[0], fmt).date()
                except ValueError:
                    continue
            return None
        except Exception:
            return None

    @action(detail=False, methods=['POST'])
    def import_excel(self, request):
        """
        Import data from Excel file with improved validation and error handling
        """
        if 'file' not in request.FILES:
            return Response({'error': 'No file uploaded'}, status=status.HTTP_400_BAD_REQUEST)
        
        excel_file = request.FILES['file']
        
        # Validate file type
        if not excel_file.name.endswith(('.xlsx', '.xls')):
            return Response({'error': 'File must be an Excel file (.xlsx or .xls)'}, 
                          status=status.HTTP_400_BAD_REQUEST)

        # Save uploaded file
        uploaded_file = UploadedExcel(file=excel_file)
        uploaded_file.save()

        try:
            dataset = pd.read_excel(excel_file)
            for _, row in dataset.iterrows():
                ExcelData.objects.create(
                    job_title=row.get('job_title', ''),
                    date_of_application=row.get('date_of_application'),
                    name=row.get('name', ''),
                    email_id=row.get('email_id', ''),
                    phone_number=str(row.get('phone_number', '')),
                    current_location=row.get('current_location', ''),
                    total_experience=str(row.get('total_experience', '')),
                    current_company_name=row.get('current_company_name', '')
                )

            return JsonResponse({
                'message': 'Excel file processed successfully',
                'rows_imported': len(dataset)
            })
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    @action(detail=False, methods=['GET'])
    def export_excel(self, request):
        ids = request.GET.get('ids')
        if ids:
            id_list = [int(id) for id in ids.split(',')]
            queryset = ExcelData.objects.filter(id__in=id_list)
        else:
            queryset = self.get_queryset()

        resource = ExcelDataResource()
        dataset = resource.export(queryset)
        response = HttpResponse(dataset.xlsx, content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="selected_data.xlsx"'
        return response

    @action(detail=False, methods=['POST'])
    def import_zip_excel(self, request):
        if 'file' not in request.FILES:
            return JsonResponse({'error': 'No file uploaded'}, status=400)

        zip_file = request.FILES['file']
        if not zip_file.name.endswith('.zip'):
            return JsonResponse({'error': 'File must be a ZIP archive'}, status=400)

        try:
            # Create a temporary directory to extract files
            with tempfile.TemporaryDirectory() as temp_dir:
                # Save the uploaded zip file
                temp_zip_path = os.path.join(temp_dir, 'upload.zip')
                with open(temp_zip_path, 'wb') as f:
                    for chunk in zip_file.chunks():
                        f.write(chunk)

                # Extract the zip file
                with zipfile.ZipFile(temp_zip_path, 'r') as zip_ref:
                    zip_ref.extractall(temp_dir)

                # Process each Excel file in the extracted directory
                successful_imports = 0
                failed_imports = 0

                for root, dirs, files in os.walk(temp_dir):
                    for file in files:
                        if file.endswith(('.xlsx', '.xls')):
                            file_path = os.path.join(root, file)
                            try:
                                # Read Excel file
                                df = pd.read_excel(file_path)
                                
                                # Process each row
                                for _, row in df.iterrows():
                                    ExcelData.objects.create(
                                        job_title=row.get('job_title', ''),
                                        date_of_application=row.get('date_of_application'),
                                        name=row.get('name', ''),
                                        email_id=row.get('email_id', ''),
                                        phone_number=row.get('phone_number', ''),
                                        current_location=row.get('current_location', ''),
                                        preferred_locations=row.get('preferred_locations', ''),
                                        total_experience=row.get('total_experience', ''),
                                        current_company_name=row.get('current_company_name', '')
                                    )
                                successful_imports += 1
                            except Exception as e:
                                failed_imports += 1
                                print(f"Error processing {file}: {str(e)}")

                return JsonResponse({
                    'message': 'ZIP file processed successfully',
                    'successful_imports': successful_imports,
                    'failed_imports': failed_imports
                })

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    @action(detail=False, methods=['GET'])
    def task_status(self, request):
        """
        Check the status of a running import task
        """
        task_id = request.query_params.get('task_id')
        if not task_id:
            return Response({'error': 'No task ID provided'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            task = process_zip_file.AsyncResult(task_id)
            
            if task.state == 'PENDING':
                response = {
                    'state': task.state,
                    'status': {
                        'message': 'Task is pending...',
                        'progress': 0,
                        'current_file': None,
                        'total_processed': 0
                    }
                }
            elif task.state == 'PROGRESS':
                info = task.info or {}
                response = {
                    'state': task.state,
                    'status': {
                        'progress': info.get('progress', 0),
                        'current_file': info.get('current_file', ''),
                        'total_processed': info.get('total_processed', 0),
                        'total_files': info.get('total_files', 0),
                        'processed_files': info.get('processed_files', 0)
                    }
                }
            elif task.state == 'SUCCESS':
                result = task.get()  # Get the actual result
                response = {
                    'state': task.state,
                    'status': {
                        'message': 'Task completed successfully',
                        'total_processed': result.get('total_processed', 0),
                        'total_files': result.get('total_files', 0),
                        'successful_files': result.get('successful_files', 0),
                        'failed_files': result.get('failed_files', 0),
                        'results': result.get('results', [])
                    }
                }
            elif task.state == 'FAILURE':
                response = {
                    'state': task.state,
                    'status': {
                        'message': 'Task failed',
                        'error': str(task.info) if task.info else 'Unknown error occurred'
                    }
                }
            else:
                response = {
                    'state': task.state,
                    'status': {
                        'message': f'Task is in {task.state} state',
                        'info': str(task.info) if task.info else None
                    }
                }
            
            return Response(response)
            
        except Exception as e:
            logger.error(f"Error checking task status: {str(e)}")
            return Response({
                'state': 'ERROR',
                'status': {
                    'message': 'Error checking task status',
                    'error': str(e)
                }
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class DownloadHistoryListView(ListView):
    model = DownloadHistory
    template_name = 'aap_api/download_history.html'
    context_object_name = 'download_history'
    paginate_by = 20  # Adjust as needed

    def get_queryset(self):
        return DownloadHistory.objects.order_by('-download_time')

def upload_excel(request):
    if request.method == 'POST' and request.FILES.get('file'):
        file = request.FILES['file']
        file_extension = os.path.splitext(file.name)[1].lower()

        try:
            if file_extension == '.zip':
                # Handle ZIP file
                zip_data = io.BytesIO(file.read())
                files_data = []

                with zipfile.ZipFile(zip_data) as zip_archive:
                    excel_files = [f for f in zip_archive.namelist() if f.lower().endswith(('.xlsx', '.xls'))]

                    for excel_file in excel_files:
                        excel_data = zip_archive.read(excel_file)
                        wb = openpyxl.load_workbook(io.BytesIO(excel_data))
                        ws = wb.active
                        data = []

                        headers = [cell.value for cell in ws[1]]
                        for row in ws.iter_rows(min_row=2, values_only=True):
                            row_data = dict(zip(headers, row))
                            data.append(row_data)

                        files_data.append({
                            'name': excel_file,
                            'data': data[:100]  # Limit preview to 100 rows
                        })

                return JsonResponse({'files': files_data})

            elif file_extension in ['.xlsx', '.xls']:
                # Handle Excel file
                wb = openpyxl.load_workbook(file)
                ws = wb.active
                data = []

                headers = [cell.value for cell in ws[1]]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    row_data = dict(zip(headers, row))
                    data.append(row_data)

                return JsonResponse({
                    'files': [{
                        'name': file.name,
                        'data': data[:100]  # Limit preview to 100 rows
                    }]
                })
            else:
                return JsonResponse({'error': 'Invalid file format'}, status=400)

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    return render(request, 'aap_api/upload.html')
